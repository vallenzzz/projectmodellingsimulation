import simpy
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─── Parameters ───────────────────────────────────────────────────────────────
N_CUSTOMERS       = 300
N_SERVERS         = 3
RANDOM_SEED       = 42
INTERARRIVAL_TIME = 30.0   # minutes (exponential mean)
SERVICE_TIME      = 40.0   # minutes (exponential mean)

SERVER_COST_PER_MIN  = 0.50   # RM per server per minute
WAIT_PENALTY_PER_MIN = 0.20   # RM per minute a customer waits

# ─── Shared data stores ───────────────────────────────────────────────────────
event_log       = []
queue_timeline  = []
server_timeline = []

def log_timeline(env, airport):
    q_len = len(airport.queue)
    busy  = airport.count
    queue_timeline.append((round(env.now, 4), q_len))
    server_timeline.append((round(env.now, 4), busy))

# ─── SimPy processes ──────────────────────────────────────────────────────────
def customer(env, name, cid, airport):
    arrival_time = env.now
    log_timeline(env, airport)

    with airport.request() as req:
        yield req
        service_start = env.now
        wait_time     = service_start - arrival_time
        log_timeline(env, airport)

        service_duration = np.random.exponential(SERVICE_TIME)
        yield env.timeout(service_duration)

    departure_time = env.now
    log_timeline(env, airport)

    event_log.append({
        "Customer ID"               : cid,
        "Arrival Time (min)"        : round(arrival_time, 4),
        "Service Start (min)"       : round(service_start, 4),
        "Wait Time (min)"           : round(wait_time, 4),
        "Service Duration (min)"    : round(service_duration, 4),
        "Departure Time (min)"      : round(departure_time, 4),
        "Total Time in System (min)": round(departure_time - arrival_time, 4),
    })

def customer_generator(env, airport):
    for cid in range(1, N_CUSTOMERS + 1):
        yield env.timeout(np.random.exponential(INTERARRIVAL_TIME))
        env.process(customer(env, f"Customer {cid}", cid, airport))

# ─── Run simulation ───────────────────────────────────────────────────────────
print("---Airport Simulation Starting ---")
np.random.seed(RANDOM_SEED)
env     = simpy.Environment()
airport = simpy.Resource(env, capacity=N_SERVERS)
env.process(customer_generator(env, airport))
env.run()
print("--- Simulation Complete ---\n")

# ─── Compute summary metrics ──────────────────────────────────────────────────
wait_times   = [r["Wait Time (min)"] for r in event_log]
total_times  = [r["Total Time in System (min)"] for r in event_log]
sim_duration = max(r["Departure Time (min)"] for r in event_log)

avg_wait      = np.mean(wait_times)
max_wait      = np.max(wait_times)
avg_total     = np.mean(total_times)
throughput_hr = len(event_log) / (sim_duration / 60)
peak_queue    = max(q for _, q in queue_timeline)
arrival_rate  = N_CUSTOMERS / sim_duration
avg_queue_len = arrival_rate * avg_wait
total_service = sum(r["Service Duration (min)"] for r in event_log)
utilisation   = total_service / (N_SERVERS * sim_duration) * 100
server_cost   = N_SERVERS * SERVER_COST_PER_MIN * sim_duration
wait_cost     = sum(wait_times) * WAIT_PENALTY_PER_MIN
total_cost    = server_cost + wait_cost

print(f"Total passengers served : {len(event_log)}")
print(f"Simulation duration     : {sim_duration:.2f} min ({sim_duration/60:.2f} hrs)")
print(f"Average wait time       : {avg_wait:.2f} min")
print(f"Maximum wait time       : {max_wait:.2f} min")
print(f"Server utilisation      : {utilisation:.2f}%")
print(f"Throughput              : {throughput_hr:.2f} customers/hr")
print(f"Total system cost       : RM {total_cost:.2f}")

# ─── Graphs ───────────────────────────────────────────────────────────────────
queue_hours  = [t / 60 for t, _ in queue_timeline]
queue_lens   = [q for _, q in queue_timeline]
server_hours = [t / 60 for t, _ in server_timeline]
server_busy  = [s for _, s in server_timeline]

plt.figure(figsize=(12, 5))
plt.step(queue_hours, queue_lens, where='post', color='steelblue', linewidth=1.2)
plt.fill_between(queue_hours, queue_lens, step='post', alpha=0.15, color='steelblue')
plt.title("Number of Passengers in Queue vs. Time", fontsize=13, fontweight='bold')
plt.xlabel("Time (Hours)")
plt.ylabel("Customers Waiting in Queue")
plt.ylim(bottom=0)
plt.xlim(left=0)
plt.grid(True, alpha=0.4)
plt.tight_layout()
plt.show()

plt.figure(figsize=(12, 5))
plt.step(server_hours, server_busy, where='post', color='seagreen', linewidth=1.2)
plt.fill_between(server_hours, server_busy, step='post', alpha=0.15, color='seagreen')
plt.axhline(y=N_SERVERS, color='red', linestyle='--', linewidth=1, label=f'Max servers ({N_SERVERS})')
plt.title("Number of Busy Servers vs. Time", fontsize=13, fontweight='bold')
plt.xlabel("Time (Hours)")
plt.ylabel("Busy Servers")
plt.ylim(bottom=0, top=N_SERVERS + 1)
plt.xlim(left=0)
plt.legend()
plt.grid(True, alpha=0.4)
plt.tight_layout()
plt.show()

# ─── Build Excel workbook ─────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Style helpers
BLUE_DARK  = "1F4E79"
BLUE_MID   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
BLUE_ALT   = "EBF3FB"
WHITE      = "FFFFFF"
YELLOW_HL  = "FFF2CC"

thin = Side(style="thin", color="AAAAAA")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_cell(ws, row, col, value, bg=BLUE_DARK, fg=WHITE, bold=True, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=fg, size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = box
    return c

def data_cell(ws, row, col, value, bg=WHITE, number_format=None, bold=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=10, bold=bold)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = box
    if number_format:
        c.number_format = number_format
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 1: Event Log ────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Event Log"

ws1.merge_cells("A1:H1")
c = ws1.cell(row=1, column=1, value="AIRPORT SIMULATION — FULL EVENT LOG (300 Customers)")
c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

ws1.merge_cells("A2:H2")
c2 = ws1.cell(row=2, column=1,
    value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
          f"Seed: {RANDOM_SEED}  |  Servers: {N_SERVERS}  |  "
          f"Mean IAT: {INTERARRIVAL_TIME} min  |  Mean Service: {SERVICE_TIME} min")
c2.font      = Font(name="Arial", size=9, color="555555")
c2.fill      = PatternFill("solid", fgColor=BLUE_LIGHT)
c2.alignment = Alignment(horizontal="center")

headers = [
    "Customer ID", "Arrival Time\n(min)", "Service Start\n(min)",
    "Wait Time\n(min)", "Service Duration\n(min)", "Departure Time\n(min)",
    "Total Time in\nSystem (min)", "Waited?"
]
for col, h in enumerate(headers, start=1):
    hdr_cell(ws1, 3, col, h, wrap=True)
ws1.row_dimensions[3].height = 30

for i, row in enumerate(event_log):
    r      = i + 4
    bg     = BLUE_ALT if i % 2 == 0 else WHITE
    waited = row["Wait Time (min)"] > 0

    data_cell(ws1, r, 1, row["Customer ID"],                  bg=bg, align="center")
    data_cell(ws1, r, 2, row["Arrival Time (min)"],           bg=bg, number_format="0.0000")
    data_cell(ws1, r, 3, row["Service Start (min)"],          bg=bg, number_format="0.0000")

    c_wait = data_cell(ws1, r, 4, row["Wait Time (min)"],
                       bg=YELLOW_HL if waited else bg, number_format="0.0000")
    if waited:
        c_wait.font = Font(name="Arial", size=10, color="C00000")

    data_cell(ws1, r, 5, row["Service Duration (min)"],       bg=bg, number_format="0.0000")
    data_cell(ws1, r, 6, row["Departure Time (min)"],         bg=bg, number_format="0.0000")
    data_cell(ws1, r, 7, row["Total Time in System (min)"],   bg=bg, number_format="0.0000")
    data_cell(ws1, r, 8, "YES" if waited else "no",
              bg=YELLOW_HL if waited else bg, bold=waited, align="center")

# Totals row
tot_row = N_CUSTOMERS + 4
hdr_cell(ws1, tot_row, 1, "TOTALS / AVERAGES", bg=BLUE_MID, align="left")
formulas = [
    f"=AVERAGE(B4:B{tot_row-1})",
    f"=AVERAGE(C4:C{tot_row-1})",
    f"=AVERAGE(D4:D{tot_row-1})",
    f"=AVERAGE(E4:E{tot_row-1})",
    f"=MAX(F4:F{tot_row-1})",
    f"=AVERAGE(G4:G{tot_row-1})",
    f'=COUNTIF(H4:H{tot_row-1},"YES")',
]
for ci, formula in enumerate(formulas, start=2):
    c = ws1.cell(tot_row, ci, value=formula)
    c.font      = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=BLUE_MID)
    c.border    = box
    c.alignment = Alignment(horizontal="center", vertical="center")
    if ci < 9:
        c.number_format = "0.0000"

set_col_widths(ws1, [13, 16, 18, 16, 20, 18, 22, 10])
ws1.freeze_panes = "A4"

# ── Sheet 2: Timeline ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Timeline")

ws2.merge_cells("A1:D1")
c = ws2.cell(row=1, column=1, value="SIMULATION TIMELINE — Queue & Server State at Every Event")
c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

for col, h in enumerate(["Time (min)", "Time (hours)", "Queue Length", "Busy Servers"], start=1):
    hdr_cell(ws2, 2, col, h)

time_map = {}
for t, q in queue_timeline:
    time_map.setdefault(t, {})["queue"] = q
for t, s in server_timeline:
    time_map.setdefault(t, {})["servers"] = s

for i, t in enumerate(sorted(time_map.keys())):
    r  = i + 3
    bg = BLUE_ALT if i % 2 == 0 else WHITE
    data_cell(ws2, r, 1, t,                               bg=bg, number_format="0.0000")
    data_cell(ws2, r, 2, round(t / 60, 6),                bg=bg, number_format="0.000000")
    data_cell(ws2, r, 3, time_map[t].get("queue", ""),    bg=bg)
    data_cell(ws2, r, 4, time_map[t].get("servers", ""),  bg=bg)

set_col_widths(ws2, [14, 14, 14, 14])
ws2.freeze_panes = "A3"

# ── Sheet 3: Summary & Cost ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Summary & Cost")

ws3.merge_cells("A1:C1")
c = ws3.cell(row=1, column=1, value="SIMULATION SUMMARY — Base Model Metrics & Cost Analysis")
c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

sections = [
    ("OPERATIONAL METRICS", [
        ("Total Customers Served",    N_CUSTOMERS,               "0"),
        ("Simulation Duration (min)", round(sim_duration, 2),    "0.00"),
        ("Simulation Duration (hrs)", round(sim_duration/60, 2), "0.00"),
        ("Server Utilisation (%)",    round(utilisation, 2),     "0.00"),
        ("Probability Servers Idle (%)", round(100-utilisation, 2), "0.00"),
        ("Throughput (customers/hr)", round(throughput_hr, 2),   "0.00"),
    ]),
    ("WAITING & QUEUE METRICS", [
        ("Average Wait Time (min)",   round(avg_wait, 4),        "0.0000"),
        ("Maximum Wait Time (min)",   round(max_wait, 4),        "0.0000"),
        ("Avg Queue Length (Lq)",     round(avg_queue_len, 4),   "0.0000"),
        ("Peak Queue Length",         peak_queue,                "0"),
        ("Avg Total Time in System",  round(avg_total, 4),       "0.0000"),
    ]),
    ("COST ANALYSIS", [
        ("Server Cost/min (each)",    SERVER_COST_PER_MIN,       "0.00"),
        ("Wait Penalty/min",          WAIT_PENALTY_PER_MIN,      "0.00"),
        ("Server Operating Cost (RM)",round(server_cost, 2),     "0.00"),
        ("Customer Waiting Cost (RM)",round(wait_cost, 2),       "0.00"),
        ("TOTAL SYSTEM COST (RM)",    round(total_cost, 2),      "0.00"),
    ]),
]

current_row = 3
for section_title, rows in sections:
    ws3.merge_cells(f"A{current_row}:C{current_row}")
    hdr_cell(ws3, current_row, 1, section_title, bg=BLUE_MID, align="left")
    ws3.row_dimensions[current_row].height = 20
    current_row += 1
    for i, (label, value, fmt) in enumerate(rows):
        bg   = BLUE_ALT if i % 2 == 0 else WHITE
        bold = "TOTAL" in label
        bg   = YELLOW_HL if bold else bg
        data_cell(ws3, current_row, 1, label, bg=bg, align="left", bold=bold)
        data_cell(ws3, current_row, 2, value, bg=bg, number_format=fmt, bold=bold)
        current_row += 1
    current_row += 1

set_col_widths(ws3, [32, 20, 10])

# ── Sheet 4: Parameters ───────────────────────────────────────────────────────
ws4 = wb.create_sheet("Parameters")

ws4.merge_cells("A1:C1")
c = ws4.cell(row=1, column=1, value="SIMULATION PARAMETERS & ASSUMPTIONS")
c.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
c.fill      = PatternFill("solid", fgColor=BLUE_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

params = [
    ("SIMULATION SETUP",),
    ("Parameter", "Value", "Notes"),
    ("Number of Customers",    N_CUSTOMERS,             "Fixed per assignment requirement"),
    ("Number of Servers",      N_SERVERS,               "3 parallel service counters"),
    ("Random Seed",            RANDOM_SEED,             "Ensures reproducibility"),
    ("",),
    ("ARRIVAL PROCESS",),
    ("Parameter", "Value", "Notes"),
    ("Inter-Arrival Distribution", "Exponential",       "Memoryless Poisson arrivals"),
    ("Mean Inter-Arrival Time",    f"{INTERARRIVAL_TIME} min", f"λ = {1/INTERARRIVAL_TIME:.4f} customers/min"),
    ("",),
    ("SERVICE PROCESS",),
    ("Parameter", "Value", "Notes"),
    ("Service Distribution",   "Exponential",           "Variable treatment time"),
    ("Mean Service Time",      f"{SERVICE_TIME} min",   f"μ = {1/SERVICE_TIME:.4f} customers/min per server"),
    ("Queue Discipline",       "FCFS",                  "First-Come-First-Served"),
    ("Max Queue Capacity",     "Unlimited",             "No customer abandonment"),
    ("",),
    ("COST MODEL",),
    ("Parameter", "Value", "Notes"),
    ("Server Cost per Min",    f"RM {SERVER_COST_PER_MIN}",  "Per server, per minute"),
    ("Wait Penalty per Min",   f"RM {WAIT_PENALTY_PER_MIN}", "Per minute a customer waits"),
    ("Cost Formula",           "Total = (Servers x Cost x Duration) + (Total Wait x Penalty)", ""),
]

current_row = 2
for row_data in params:
    if len(row_data) == 1:
        if row_data[0]:
            ws4.merge_cells(f"A{current_row}:C{current_row}")
            hdr_cell(ws4, current_row, 1, row_data[0], bg=BLUE_MID, align="left")
        current_row += 1
        continue
    if row_data[0] == "Parameter":
        for ci, h in enumerate(row_data, start=1):
            hdr_cell(ws4, current_row, ci, h, bg="3A6EA0")
    else:
        for ci, val in enumerate(row_data, start=1):
            bg = BLUE_ALT if current_row % 2 == 0 else WHITE
            data_cell(ws4, current_row, ci, val, bg=bg,
                      align="left" if ci in (1, 3) else "center")
    current_row += 1

set_col_widths(ws4, [28, 20, 45])

# ─── Save & Download ──────────────────────────────────────────────────────────
out_path = "Airport_Simulation_Timestamps.xlsx"
wb.save(out_path)
print(f"\nExcel file saved: {out_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Event log rows: {len(event_log)}")
print(f"Timeline rows : {len(time_map)}")